"""Output generation: master timetable, section/teacher/room views,
conflict report CSV, validation report text, and Excel/CSV exports."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import config
from src.models import Assignment, TimetableData
from src.postvalidator import PostValidationReport
from src.solver import SolveResult
from src.validator import ValidationReport
from src.utils import get_logger

log = get_logger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

MASTER_COLS = ["Day", "Time Slot", "Room", "Course", "Course Code",
               "Section(s)", "Teacher", "Duration"]


def export_all(data: TimetableData, result: SolveResult,
               input_report: ValidationReport,
               post_report: PostValidationReport,
               out_dir: Path | None = None) -> None:
    """Generate every required output artifact."""
    out_dir = out_dir or config.OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    rows = _master_rows(data, result.assignments)

    _write_csv(out_dir / config.OUTPUT_TIMETABLE_CSV.name, rows)
    _write_excel(out_dir / config.OUTPUT_TIMETABLE_XLSX.name, data,
                 result, rows)
    _write_conflict_report(out_dir / config.OUTPUT_CONFLICT_REPORT.name,
                           input_report, post_report)
    _write_validation_report(out_dir / config.OUTPUT_VALIDATION_REPORT.name,
                             result, post_report)
    log.info("All outputs written to %s", out_dir)


def _master_rows(data: TimetableData,
                 assignments: list[Assignment]) -> list[list[str]]:
    session_map = {a.session.session_id: a for a in assignments}
    rows = []
    for session in data.sessions:
        a = session_map.get(session.session_id)
        rows.append([
            a.day if a else "",
            a.period_label if a else "",
            a.room if a else "",
            session.course.title.strip(),
            session.course.code,
            session.course.section,
            session.teacher or "TBA",
            "90 min",
        ])
    # stable ordering: day, period, room
    day_order = {d: i for i, d in enumerate(config.VALID_DAYS)}
    rows.sort(key=lambda r: (day_order.get(r[0], 9), r[1], r[2]))
    return rows


def _write_csv(path: Path, rows: list[list[str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(MASTER_COLS)
        w.writerows(rows)
    log.info("CSV written: %s (%d rows)", path, len(rows))


def _write_excel(path: Path, data: TimetableData, result: SolveResult,
                 rows: list[list[str]]) -> None:
    wb = Workbook()

    # ---- Sheet 1: master timetable ----
    ws = wb.active
    ws.title = "Master Timetable"
    _style_sheet_header(ws, MASTER_COLS)
    for i, r in enumerate(rows, start=2):
        for j, v in enumerate(r, start=1):
            cell = ws.cell(i, j, v)
            cell.alignment = WRAP
            if i % 2 == 0:
                cell.fill = ALT_FILL
    _set_widths(ws, [10, 14, 12, 38, 12, 16, 22, 10])

    # ---- Sheet 2: section-wise timetable (grid style) ----
    ws2 = wb.create_sheet("Section-Wise")
    headers = ["Section"] + [f"{d} {config.PERIODS[p]}"
                             for d in config.VALID_DAYS
                             for p in config.ACTIVE_PERIODS]
    _style_sheet_header(ws2, headers)
    sessions_of = {s.session_id: s for s in data.sessions}
    section_names = sorted(data.section_sessions.keys())
    for i, sec in enumerate(section_names, start=2):
        ws2.cell(i, 1, sec)
        placed = 0
        for a in result.assignments:
            if sec in a.session.sections:
                cell_txt = (f"{a.session.course.title}\n"
                            f"R:{a.room}\n"
                            f"T:{a.session.teacher or 'TBA'}")
                placed += 1
                col = 2 + (config.ACTIVE_PERIODS.index(a.period)
                           + config.VALID_DAYS.index(a.day)
                           * len(config.ACTIVE_PERIODS))
                prev = ws2.cell(i, col).value
                ws2.cell(i, col, (prev + "\n---\n" if prev else "") + cell_txt)
        if placed:
            ws2.cell(i, 1).font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22
    ws2.row_dimensions[1].height = 30

    # ---- Sheet 3: teacher-wise timetable ----
    ws3 = wb.create_sheet("Teacher-Wise")
    _style_sheet_header(ws3, headers)
    teachers = sorted(t for t in data.teacher_sessions if t)
    for i, t in enumerate(teachers, start=2):
        ws3.cell(i, 1, t)
        for a in result.assignments:
            if a.session.teacher != t:
                continue
            cell_txt = (f"{a.session.course.title} ({a.session.course.section})\n"
                        f"R:{a.room}")
            col = 2 + (config.ACTIVE_PERIODS.index(a.period)
                       + config.VALID_DAYS.index(a.day)
                       * len(config.ACTIVE_PERIODS))
            prev = ws3.cell(i, col).value
            ws3.cell(i, col, (prev + "\n---\n" if prev else "") + cell_txt)
        ws3.cell(i, 1).font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # ---- Sheet 4: room-wise timetable ----
    ws4 = wb.create_sheet("Room-Wise")
    _style_sheet_header(ws4, headers)
    for i, room in enumerate(sorted(data.rooms), start=2):
        ws4.cell(i, 1, room)
        for a in result.assignments:
            if a.room != room:
                continue
            cell_txt = (f"{a.session.course.title} ({a.session.course.section})\n"
                        f"T:{a.session.teacher or 'TBA'}")
            col = 2 + (config.ACTIVE_PERIODS.index(a.period)
                       + config.VALID_DAYS.index(a.day)
                       * len(config.ACTIVE_PERIODS))
            prev = ws4.cell(i, col).value
            ws4.cell(i, col, (prev + "\n---\n" if prev else "") + cell_txt)
        ws4.cell(i, 1).font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 22

    # ---- Sheet 5: solver status ----
    ws5 = wb.create_sheet("Solver Status")
    ws5["A1"] = "Solver status"
    ws5["B1"] = result.status
    ws5["A2"] = "Objective value (penalties)"
    ws5["B2"] = result.objective_value
    ws5["A3"] = "Wall time (s)"
    ws5["B3"] = round(result.wall_time_seconds, 2)
    ws5["A4"] = "Note"
    ws5["B4"] = result.message
    ws5["B4"].alignment = WRAP
    for row in range(1, 5):
        ws5.cell(row, 1).font = Font(bold=True)
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 90

    wb.save(path)
    log.info("Excel written: %s", path)


def _style_sheet_header(ws, headers: list[str]) -> None:
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(1, j, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.freeze_panes = "C2"


def _set_widths(ws, widths: list[float]) -> None:
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _write_conflict_report(path: Path, input_report: ValidationReport,
                           post_report: PostValidationReport) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["Stage", "Kind", "Day", "Period", "Room", "Detail",
                    "Involved"])
        for c in input_report.conflicts:
            w.writerow(["INPUT", c.kind, c.day or "",
                        config.PERIODS[c.period] if c.period is not None else "",
                        c.room or "", c.detail, " | ".join(c.items)])
        if not input_report.conflicts:
            w.writerow(["INPUT", "none", "", "", "",
                        "No input conflicts detected", ""])
        w.writerow([])
        w.writerow(["Stage", "Kind", "Detail"])
        if post_report.violations:
            for v in post_report.violations:
                w.writerow(["POST-SOLUTION", "violation", v])
        else:
            w.writerow(["POST-SOLUTION", "none",
                        "All hard constraints satisfied (VALIDATION PASSED)"])
    log.info("Conflict report written: %s", path)


def _write_validation_report(path: Path, result: SolveResult,
                             post_report: PostValidationReport) -> None:
    lines = [
        "=" * 70,
        "VALIDATION REPORT - AI-Assisted Timetable Generator",
        "=" * 70,
        "",
        "SOLVER RESULT",
        f"  Status          : {result.status}",
        f"  Objective value : {result.objective_value}",
        f"  Wall time       : {result.wall_time_seconds:.2f} s",
        f"  Branches        : {result.branches}",
        f"  Conflicts found : {result.conflicts}",
        f"  Note            : {result.message}",
        "",
        "ASSIGNMENTS",
        f"  Total sessions scheduled : {len(result.assignments)}",
        "",
    ] + post_report.to_text().split("\n")
    path.write_text("\n".join(lines), encoding="utf-8")
    log.info("Validation report written: %s", path)
