"""Excel data loading and preprocessing pipeline.

Responsibilities:
    * Read every relevant sheet of the workbook with openpyxl (merged cells
      handled automatically: openpyxl back-fills values into all cells of a
      merged range when ``data_only=True`` is used, so grid cells that inherit
      a day/room label are read correctly).
    * Clean empty rows and irrelevant headers.
    * Normalize day names, time-slot formats, teacher and room names.
    * Parse course information and expand course rows into scheduling sessions.
    * Detect missing values and flag them instead of silently inventing data.
"""

from __future__ import annotations

import re
from typing import Final

from src import config
from src.models import (
    Course,
    ExistingEntry,
    Session,
    TimetableData,
)
from src.utils import get_logger, is_lab_title, is_lab_room, normalize_teacher

log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Header row detection
# ---------------------------------------------------------------------------

_HEADER_TOKENS: Final[set[str]] = {"code", "course title", "section",
                                   "instructor name", "credit hours"}


def _find_header_row(ws) -> int:
    """Locate the header row of a course sheet (default row 2).

    Scans the first 10 rows and picks the row whose cells most closely match
    the known header tokens, so the loader is not hardcoded to row 2.
    """
    best, best_score = 2, 0
    for r in range(1, 11):
        hits = 0
        for c in range(1, 8):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() in _HEADER_TOKENS:
                hits += 1
        if hits > best_score:
            best, best_score = r, hits
    return best


def _load_course_sheets(wb, data: TimetableData) -> None:
    """Extract and validate course requirement rows from the program sheets."""
    for sheet in config.COURSE_SHEETS:
        if sheet not in wb.sheetnames:
            log.warning("Expected course sheet '%s' not found; skipping.", sheet)
            continue
        ws = wb[sheet]
        header_row = _find_header_row(ws)
        program = ""
        n_courses = 0
        for r in range(header_row + 1, ws.max_row + 1):
            code = ws.cell(r, config.COL_CODE).value
            title = ws.cell(r, config.COL_TITLE).value
            section = ws.cell(r, config.COL_SECTION).value
            instructor = ws.cell(r, config.COL_INSTRUCTOR).value
            credit = ws.cell(r, config.COL_CREDIT_HOURS).value
            program_cell = ws.cell(r, config.COL_PROGRAM).value

            # Rows without a course code are group headers / category notes
            # (e.g. 'BS(SE)-2026', 'Labs', 'Arts and Humanities Elective',
            # 'Repeat Courses', 'Merged with DS'). They set the batch context
            # but are NOT courses themselves and must not be scheduled.
            if not code:
                if title:
                    title_stripped = str(title).strip()
                    if (title_stripped.startswith("BS(")
                            or title_stripped.startswith("MS(")
                            or re.search(r"batch", title_stripped,
                                         re.IGNORECASE)):
                        program = title_stripped
                continue

            # Skip fully empty or junk rows.
            if not code and not title and not section:
                continue
            title_s = str(title).strip() if title else ""
            if not title_s:
                continue

            code_s = str(code).strip() if code else ""
            section_s = str(section).strip() if section else ""
            inst_s = normalize_teacher(str(instructor)) if instructor else ""
            program_s = str(program_cell).strip() if program_cell else program

            lab = is_lab_title(title_s)
            merged = bool(inst_s and config.MERGED_INSTRUCTOR_TOKEN in inst_s.lower())

            ch = _parse_credit_hours_cell(credit, title_s, code_s, sheet, r)
            if ch is None:
                # Row exists but credit hours unparseable -> 1 session, flagged.
                ch = 1
                data.unparseable_credit_hours.append(
                    Course(code_s, title_s, section_s, inst_s, 1, program_s,
                           str(ws.cell(r, config.COL_CATEGORY).value or ""),
                           sheet, lab, merged))
                log.warning("Unparseable credit hours at %s!R%d (%s %s); using 1.",
                            sheet, r, code_s, title_s)

            course = Course(code_s, title_s, section_s, inst_s, ch, program_s,
                            str(ws.cell(r, config.COL_CATEGORY).value or ""),
                            sheet, lab, merged)

            if merged:
                data.merged_courses.append(course)
                log.info("Merged course flagged: %s %s (%s)", code_s, title_s,
                         section_s)
                continue

            if lab:
                # Lab rows: one session per row regardless of credit hours.
                # Each half-section (e.g. BAI-1A1, BAI-1A2) gets its own
                # session; combined lab rows (BCS-1A/3A) create one session
                # covering both sections.
                n_sessions = 1
            else:
                n_sessions = course.credit_hours

            for k in range(1, n_sessions + 1):
                sid = f"S-{len(data.sessions) + 1:04d}"
                data.sessions.append(
                    Session(sid, course, k, lab))

            data.courses.append(course)
            n_courses += 1

            if not inst_s:
                data.missing_teachers.append(course)
            elif merged:
                pass  # already counted above

        log.info("Sheet '%s': %d course rows, %d sessions created.",
                 sheet, n_courses,
                 sum(1 for s in data.sessions if s.course.sheet == sheet))


def _parse_credit_hours_cell(value, title: str, code: str, sheet: str,
                             row: int):
    from src.utils import parse_credit_hours
    return parse_credit_hours(value)


def _load_grid_sheet(wb, data: TimetableData) -> None:
    """Parse the master timetable grid (day x room x period cells).

    The grid layout (verified against the workbook):
        Row 3: 'Periods' + one period string per period block (e.g. 08:30-10:00)
        Row 4: 'Days' | 'Room' | sub-slot numbers (10..60 per block)
        Col A: day name (inherited downward until a new day appears)
        Col B: room name (inherited downward until a new room appears)
        Other cells: 'Course (SECTIONS): Teacher' or empty.

    Rooms listed in column B are collected as the global room set. Room names
    that follow the lab naming convention are additionally flagged as lab rooms.
    """
    if config.GRID_SHEET not in wb.sheetnames:
        log.warning("Grid sheet '%s' not found; no existing assignments loaded.",
                    config.GRID_SHEET)
        return
    ws = wb[config.GRID_SHEET]

    # --- resolve period start columns from the period header row -------------
    period_starts: list[tuple[int, int, str]] = []  # (start_col, end_col, label)
    for c in range(config.GRID_FIRST_CONTENT_COL, ws.max_column + 1):
        v = ws.cell(config.GRID_PERIOD_HEADER_ROW, c).value
        if isinstance(v, str):
            label = v.strip()
            if re.fullmatch(r"\d{2}:\d{2}-\d{2}:\d{2}", label):
                start = c
                end = ws.max_column
                if period_starts:
                    end = period_starts[-1][1]  # end before next period block
                    period_starts[-1] = (period_starts[-1][0], start - 1,
                                         period_starts[-1][2])
                period_starts.append((start, end, label))
    if not period_starts:
        log.warning("No period headers found on row %d of '%s'; using config.",
                    config.GRID_PERIOD_HEADER_ROW, config.GRID_SHEET)
        start = config.GRID_FIRST_CONTENT_COL
        for i, label in enumerate(config.PERIODS):
            end = (config.PERIOD_START_COLUMNS[i + 1] - 1
                   if i + 1 < len(config.PERIOD_START_COLUMNS)
                   else ws.max_column)
            period_starts.append((start, end, label))
            start = end + 1

    col_to_period = {}
    for start, end, label in period_starts:
        idx = config.PERIODS.index(label) if label in config.PERIODS else -1
        for c in range(start, end + 1):
            col_to_period[c] = idx

    rooms_seen: set[str] = set()
    day = ""
    room = ""
    n_entries = 0

    for r in range(config.GRID_PERIOD_HEADER_ROW + 2, ws.max_row + 1):
        d = ws.cell(r, config.GRID_DAY_COL).value
        rm = ws.cell(r, config.GRID_ROOM_COL).value
        if d and str(d).strip():
            day = str(d).strip()
        if rm and str(rm).strip():
            room = str(rm).strip()
        if not day or not room:
            continue

        rooms_seen.add(room)
        for c in range(config.GRID_FIRST_CONTENT_COL, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not v or not str(v).strip():
                continue
            txt = str(v).strip()
            period = col_to_period.get(c, -1)
            if period < 0:
                log.debug("Content cell %s outside known periods; skipped.",
                          ws.cell(r, c).coordinate)
                continue
            entry = _parse_grid_cell(day, room, period, txt)
            if entry:
                data.existing_entries.append(entry)
                n_entries += 1

    data.rooms = sorted(rooms_seen)
    data.lab_rooms = sorted(rm for rm in rooms_seen if is_lab_room(rm))
    log.info("Grid '%s': %d entries, %d rooms (%d lab rooms).",
             config.GRID_SHEET, n_entries, len(data.rooms), len(data.lab_rooms))


def _expand_section_list(raw: str) -> tuple[str, ...]:
    """Expand shorthand section lists such as ``BCS-1E/3E`` into full names.

    Tokens without an explicit prefix inherit the prefix (program code and
    dash) of the nearest preceding full token, so ``BCS-1E/3E`` becomes
    ``('BCS-1E', 'BCS-3E')`` and ``BAI-2A/2B/3A`` becomes
    ``('BAI-2A', 'BAI-2B', 'BAI-3A')``.
    """
    tokens = [t.strip() for t in raw.split("/") if t.strip()]
    expanded: list[str] = []
    prefix = ""
    for token in tokens:
        if "-" in token:
            prefix = token.rsplit("-", 1)[0] + "-"
            expanded.append(token)
        elif prefix:
            expanded.append(prefix + token)
        else:
            expanded.append(token)
    return tuple(expanded)


def _parse_grid_cell(day: str, room: str, period: int, text: str) -> ExistingEntry | None:
    """Parse ``'Calculus (BCS-1E/3E): Mazhar H'`` style cell text."""
    m = re.search(r"\(([^)]+)\)", text)
    if not m:
        log.debug("Grid cell without section list ignored: %s", text)
        return None
    sections = _expand_section_list(m.group(1))
    body = text.split("(")[0].strip()
    teacher = ""
    if ":" in text:
        teacher = normalize_teacher(text.split(":")[-1])
    return ExistingEntry(day, room, period, body, sections, teacher)


def _build_indexes(data: TimetableData) -> None:
    """Build teacher->sessions and section->sessions indexes used by the
    constraint model and the independent validator."""
    for session in data.sessions:
        t = session.teacher
        if t:
            data.teacher_sessions.setdefault(t, set()).add(session.session_id)
        for sec in session.sections:
            data.section_sessions.setdefault(sec, set()).add(session.session_id)


def load_timetable_data(path=None) -> TimetableData:
    """Full load -> preprocess -> index pipeline."""
    import openpyxl

    path = path or config.INPUT_WORKBOOK
    log.info("Loading workbook: %s", path)
    wb = openpyxl.load_workbook(path, data_only=True)

    data = TimetableData()
    _load_course_sheets(wb, data)
    _load_grid_sheet(wb, data)
    _build_indexes(data)

    log.info("Loaded %d courses, %d sessions (%d lecture + %d lab), "
             "%d teachers, %d sections.",
             len(data.courses), len(data.sessions),
             sum(1 for s in data.sessions if not s.is_lab),
             sum(1 for s in data.sessions if s.is_lab),
             len(data.teacher_sessions), len(data.section_sessions))
    return data
